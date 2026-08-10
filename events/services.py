from dataclasses import dataclass

from django.core.exceptions import ValidationError
from django.db import transaction
from django.utils.translation import gettext as _
from openpyxl import load_workbook

from .models import SpecialEventParticipant


EXPECTED_HEADERS = (
    "Na",
    "JINA LA MTAFITI",
    "TAASISI",
    "UTAFITI",
    "NYANJA",
)


@dataclass(frozen=True)
class ImportResult:
    created: int
    updated: int
    skipped: int
    sheets: int


def _text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalized_header(value):
    return " ".join(_text(value).upper().split()).rstrip(".")


@transaction.atomic
def import_special_event_participants(*, event, uploaded_file, user):
    """Import every matching worksheet while preserving existing QR tokens."""
    if not event.category.is_special_event:
        raise ValidationError(
            _("Select an event in the Special Event category.")
        )

    try:
        workbook = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise ValidationError(
            _("The uploaded file could not be read as an Excel workbook.")
        ) from exc

    expected = tuple(_normalized_header(item) for item in EXPECTED_HEADERS)
    created = updated = skipped = matching_sheets = 0
    seen_source_rows = set()

    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            actual = tuple(_normalized_header(value) for value in header[:5])
            if actual != expected:
                continue

            matching_sheets += 1
            source_sheet = worksheet.title.strip()
            for row_index, row in enumerate(rows, start=2):
                values = [_text(value) for value in row[:5]]
                if not any(values):
                    skipped += 1
                    continue
                source_number, full_name, institution, research_title, research_field = values
                if not source_number or not full_name:
                    raise ValidationError(
                        _("Sheet %(sheet)s row %(row)s must contain Na and JINA LA MTAFITI.")
                        % {"sheet": source_sheet, "row": row_index}
                    )
                source_key = (source_sheet.casefold(), source_number.casefold())
                if source_key in seen_source_rows:
                    raise ValidationError(
                        _("Sheet %(sheet)s contains duplicate participant number %(number)s.")
                        % {"sheet": source_sheet, "number": source_number}
                    )
                seen_source_rows.add(source_key)

                existing = SpecialEventParticipant.objects.filter(
                    event=event,
                    source_sheet=source_sheet,
                    source_number=source_number,
                ).first()
                participant, was_created = SpecialEventParticipant.objects.update_or_create(
                    event=event,
                    source_sheet=source_sheet,
                    source_number=source_number,
                    defaults={
                        "full_name": full_name,
                        "source_row_index": row_index,
                        "institution": institution,
                        "research_title": research_title,
                        "research_field": research_field,
                        "is_active": True,
                        "created_by": existing.created_by if existing else user,
                        "updated_by": user,
                    },
                )
                created += int(was_created)
                updated += int(not was_created)
    finally:
        workbook.close()

    if not matching_sheets:
        raise ValidationError(
            _("No worksheet has the required columns: Na, JINA LA MTAFITI, TAASISI, UTAFITI and NYANJA.")
        )
    if not created and not updated:
        raise ValidationError(_("No participant rows were found in the Excel file."))

    return ImportResult(
        created=created,
        updated=updated,
        skipped=skipped,
        sheets=matching_sheets,
    )
